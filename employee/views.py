from django.shortcuts import render, redirect  
from django.http import JsonResponse, HttpResponse

from .models import *
from master.models import *
#from employee.models *
import requests, json

from datetime import date, timedelta
from collections import Counter

import openpyxl
from openpyxl.styles import Font
import uuid


from pytz import timezone
from datetime import datetime as dt
from django.conf import settings

tdate = dt.now(timezone("Asia/Kolkata")).strftime('%Y-%m-%d')
yearmonth = dt.now(timezone("Asia/Kolkata")).strftime('%Y-%m')
time = dt.now(timezone("Asia/Kolkata")).strftime('%H:%M %p')


from django.contrib import messages
from django.db.models import Sum, F, Q


from rest_framework.decorators import api_view    
from rest_framework import serializers
from rest_framework.response import Response
from .serializers import *
from rest_framework.parsers import JSONParser

#from global_fun import *

"""
@api_view(["POST"])
def analytics(request):

    json_data = request.data
    month = int(json_data['month'])
    
    if "SalesEmployeeCode" in json_data:
        ##print("yes")
        
        SalesEmployeeCode = json_data['SalesEmployeeCode']
        if employee.objects.filter(SalesEmployeeCode = SalesEmployeeCode, Active = "tYES").exists():
            emp_obj =  employee.objects.get(SalesEmployeeCode=SalesEmployeeCode)
            empids = getAllReportingToIds(SalesEmployeeCode)
            
            empids = []
            
            tgt_all = Target.objects.filter(SalesPersonCode__in=empids).exclude(monthYear=yearmonth).order_by("-monthYear")[:month]
            #{"month":"3", "SalesEmployeeCode":"3"}
            amount = sum(tgt_all.values_list('amount', flat=True))            
            ##print(amount)
            #amount = "{:.2f}".format(amount)
            ###print(amount)
            
            sale = sum(tgt_all.values_list('sale', flat=True))
            ##print(sale)
            
            sale_diff = sum(tgt_all.values_list('sale_diff', flat=True))
            ##print(sale_diff)
            
            notification = Notification.objects.filter(Emp=emp_obj.id, CreatedDate=tdate, Read=0).order_by("-id").count()
            ##print(notification)
            
            
            return Response({"message": "Success","status": 200,"data":[{"notification":notification, "amount":amount, "sale":sale, "sale_diff":sale_diff}]})
            
            #return Response({"message": "Success","status": 201,"data":[{"emp":SalesEmployeeCode}]})
        else:
            return Response({"message": "Unsuccess","status": 201,"data":[{"error":"SalesEmployeeCode?"}]})
    else:
        return Response({"message": "Unsuccess","status": 201,"data":[{"error":"SalesEmployeeCode?"}]})


# employee dashboard
@api_view(["POST"])
def dashboard(request):
    json_data = request.data
    try:
        SalesEmployeeCode = json_data['SalesEmployeeCode']
        if employee.objects.filter(SalesEmployeeCode = SalesEmployeeCode, Active = "tYES").exists():

            # empObj = employee.objects.get(SalesEmployeeCode = SalesEmployeeCode)
            empList = getAllReportingToIds(SalesEmployeeCode)
            ##print(empList)
            
            # emp_ids = employee.objects.filter(SalesEmployeeCode__in=empList).values_list('id', flat=True)
            # ##print(emp_ids)
            #{"SalesEmployeeCode":4}
            
            lead_all = Lead.objects.filter(assignedTo__in=empList).count()
            ##print(lead_all)
            
            opp_all = Opportunity.objects.filter(SalesPerson__in=empList).count()
            ###print(opp_all)
            
            quot_all = Quotation.objects.filter(SalesPersonCode__in=empList).count()
            ###print(quot_all)
            
            ord_all = Order.objects.filter(SalesPersonCode__in=empList).count()
            ###print(ord_all)
            
            #bp_all = BusinessPartner.objects.filter(SalesPersonCode__in=empList).count()
            bp_all = BusinessPartner.objects.all().count()
            ###print(bp_all)
            
            tgt_all = Target.objects.filter(SalesPersonCode__in=empList, monthYear=yearmonth)
            
            amount = sum(tgt_all.values_list('amount', flat=True))            
            ##print(amount)

            sale = 0
            orderAmtList = list(Order.objects.filter(SalesPersonCode__in=empList).values_list('DocTotal', flat=True))
            for amt in orderAmtList:
                sale = sale+float(amt)
            ##print(sale)
            
            sale_diff = float(sale - float(amount))
            ##print(sale_diff)
            
            emp_objj = employee.objects.filter(SalesEmployeeCode=SalesEmployeeCode).first()
            notification = Notification.objects.filter(Emp=emp_objj.SalesEmployeeCode, CreatedDate=tdate, Read=0).order_by("-id").count()
            ##print(notification)

            ord_over = Order.objects.filter(SalesPersonCode__in=empList, DocumentStatus="bost_Open", DocDueDate__lt=tdate).count()
            ##print(ord_over)
            ##print(date)
            
            ord_open = Order.objects.filter(SalesPersonCode__in=empList, DocumentStatus="bost_Open", DocDueDate__gte=tdate).count()
            ##print(ord_open)

            ord_close = Order.objects.filter(SalesPersonCode__in=empList, DocumentStatus="bost_Close").count()
            ##print(ord_close)

            inv_all = Invoice.objects.filter(SalesPersonCode__in=empList).count()
            ###print(inv_all)
            
            tnd_all = Tender.objects.filter(SalesPersonCode__in=empList).count()
            ##print(tnd_all)
            projectCount = Project.objects.filter(project_owner__in=empList).count()
            
            #added by millan on 25/08/2022 for Campset Count
            campset_count = CampaignSet.objects.all().count()
			
    
            # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
            # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
            context = {
                "notification":notification, 
                "amount":amount, 
                "sale":sale, 
                "sale_diff":sale_diff, 
                "Opportunity":opp_all, 
                "Quotation":quot_all, 
                "Order":ord_all, 
                "Customer":bp_all, 
                "Leads":lead_all, 
                "Over":ord_over, 
                "Open":ord_open, 
                "Close":ord_close, 
                "campset_count" : campset_count,
                "Invoice":inv_all, 
                "Tender":tnd_all,
                "Project":projectCount,
            }
            return Response({"message": "Success","status": 200,"data":[context]})
        else:
            return Response({"message": "Unsuccess","status": 201,"data":[{"error":"SalesEmployeeCode?"}]})
    except Exception as e:
        return Response({"message": str(e),"status": 201,"data":[]})

        
#Employee All API
@api_view(["GET"])
def all(request):
    
    try:
        #emps_obj = employee.objects.all().exclude(SalesEmployeeCode="24").order_by('-id')
        emps_obj = employee.objects.all()
        emps_json = employee_serializer(emps_obj, many=True)
        emps_json = json.loads(json.dumps(emps_json.data))
                
        for employee_obj in emps_json:
            ##print(employee_obj['div'])
            if employee_obj['div'] !="":
                div_arr = employee_obj['div'].split(",")
                div_obj = Category.objects.filter(Number__in=div_arr) 
                div_json = CategorySerializer(div_obj, many=True).data
                employee_obj['div'] = div_json
            else:
                employee_obj['div'] = []

            if employee.objects.filter(SalesEmployeeCode = employee_obj['reportingTo']).exists():
                reportingToObj = employee.objects.filter(SalesEmployeeCode = employee_obj['reportingTo']).values("id", "SalesEmployeeCode", "SalesEmployeeName")
                reportingToJson = EmployeeSerializer(reportingToObj, many = True)
                employee_obj['reportingToDetails'] = reportingToJson.data
            else:
                employee_obj['reportingToDetails'] = []

        return Response({"message": "Success","status": 200,"data":emps_json})
    except Exception as e:
        return Response({"message": str(e),"status": 201,"data":[]})

@api_view(["POST"])
def all_filter(request):
    json_data = request.data
    try:
        
        SalesEmployeeCode = json_data['SalesEmployeeCode']
        if employee.objects.filter(SalesEmployeeCode = SalesEmployeeCode, Active = "tYES").exists():
            empids = getAllReportingToIds(SalesEmployeeCode)
            
            emps_all = employee.objects.filter(SalesEmployeeCode__in = empids, Active = "tYES").order_by('-id')
            emps_json = EmployeeSerializer(emps_all, many=True)

            emps_json = json.loads(json.dumps(emps_json.data))
            
            for employee_obj in emps_json:
                ##print(employee_obj['div'])
                if employee_obj['div'] !="":
                    div_arr = employee_obj['div'].split(",")
                    div_obj = Category.objects.filter(Number__in=div_arr) 
                    div_json = CategorySerializer(div_obj, many=True).data
                    employee_obj['div'] = div_json
                else:
                    employee_obj['div'] = []
            
            return Response({"message": "Success","status": 200,"data":emps_json})            
            #return Response({"message": "Success","status": 201,"data":[{"emp":SalesEmployeeCode}]})
        else:
            return Response({"message": "Invalid SalesEmployeeCode?","status": 201,"data":[]})
    except Exception as e:
        ##print("no")
        return Response({"message": str(e),"status": 201,"data":[]})

"""
from django.apps import apps


@api_view(['POST'])
def get_fields(request):
    try:		
        app_name = request.data['app']
        model_name = request.data['model']
        #print(app_name)
        #print(model_name)
        model = apps.get_model(app_label=app_name, model_name=model_name)
        #print(model)
        
        field_names = [field.name for field in model._meta.get_fields() if field.name != 'id']
        #print(field_names)


        return Response({"message":"successful","status":200, "data":field_names})
    except Exception as e:
        return Response({"message":str(e),"status":"201","data":[]})


@api_view(['POST'])
def get_data(request):
    try:		
        app_name = request.data['app']
        model_name = request.data['model']
        filterby = request.data['filterby']
        fields = request.data['fields']
        #print(app_name)
        #print(model_name)
        model = apps.get_model(app_label=app_name, model_name=model_name)
        model.objects.all()
        objs = model.objects.filter(**filterby).values(*fields).order_by("-id")
        return Response({"message":"successful","status":200, "data":objs})
    except Exception as e:
        return Response({"message":str(e),"status":"201","data":[]})





## for all table
@api_view(['POST'])
def model(request):
    try:		
        app_name = request.data['app']
        model_name = request.data['model']
        action = request.data['action']
        data = request.data['data']
        if action =="create":
            try:
                del data['id']
            except Exception:
                pass
                
            model = apps.get_model(app_label=app_name, model_name=model_name)
            obj = model(**data)
            obj.save()     
        elif action =="update":
            #print(data['id'])
            model = apps.get_model(app_label=app_name, model_name=model_name)
            obj = model.objects.filter(id=data['id']).update(**data)
        else:
            return Response({"message":"Action not Match","status":201, "data":[]})
        
        return Response({"message":"successful","status":200, "data":[]})
    except Exception as e:
        return Response({"message":str(e),"status":"201","data":[]})


# {"model_name":"position_master", "filterby":""}
"""
{
    "model": "position_master",
    "filterby": {"country": "japan"},
    "fields": ["region","country","position_id","position_name","baseline_total"],
    "sortby": ["baseline_total"]
}

{
    "model": "position_master",
    "filterby": {"country": "japan"},
    "fields": ["region","country","position_id","position_name","baseline_total"],
    "sortby": ["baseline_total"],
    "order": "desc"
}

"""


##select region,country,position_id,position_name,baseline_total from position_master WHERE 1=1 and country='japan' order by baseline_total desc limit 20,30;

#select region,country,position_id,position_name,baseline_total from master_position WHERE 1=1  and country in ('India', 'Nepal') order by baseline_total desc limit 10 OFFSET 20

@api_view(['POST'])
def get_list(request):
    try:		
        
        mydb = settings.CONNECT()
        mycursor = mydb.cursor(dictionary=True, buffered=True)
        #print(mycursor)

        model_name = request.data['model']
        filterby = request.data['filterby']
        fields = request.data['fields']
        if len(fields) > 0:
            fields = ",".join(fields)
        else:
            fields ="*"
        sortby = request.data['sortby']
        sortby = ",".join(sortby)

        order = request.data['order']
        
        limit = settings.LIMIT(request.data)
        filterquery = ""
        for key,value in filterby.items():
            if type(value)==list:
                value = "', '".join(value)
                filterquery = filterquery + f" and {key} in ('{value}')"
            else:
                filterquery = filterquery + f" and {key}='{value}'"
        #print("filterquery", filterquery)
        
        sql = f"select {fields} from {model_name} WHERE 1=1 {filterquery} order by {sortby} {order} {limit}"
        #sql = f"select * from {model_name} WHERE 1=1 {filterquery}"
        #print(sql)
        mycursor.execute(sql)
        objs = mycursor.fetchall()
        counts = mycursor.rowcount
        

        
        sql = f"select count(id) as count from {model_name} WHERE 1=1 {filterquery}"
        #print(sql)
        mycursor.execute(sql)
        count = mycursor.fetchone()
        #print(f'count: {count}')
        
        mycursor.close()
        
        #objs = model.objects.filter(**filterby).values(*fields).order_by("-id")
        return Response({"message":"successful","status":200, "data":objs, "counts":count})
    except Exception as e:
        return Response({"message":str(e),"status":"201","data":[]})

@api_view(['POST'])
def get_list_pos(request):
    try:		
        
        mydb = settings.CONNECT()
        mycursor = mydb.cursor(dictionary=True, buffered=True)
        #print(mycursor)

        model_name = request.data['model']
        filterby = request.data['filterby']
        if "search" in request.data:
            search = request.data['search']
            
            if len(search) >= 3:
                search = f" and (position_id like '%{search}%' or position_name like '%{search}%' or emp_ref_no like '%{search}%')"
            else:
                search =""
        else:
            search =""
        
        fields = request.data['fields']
        if len(fields) > 0:
            fields = ",".join(fields)
        else:
            fields ="*"
        sortby = request.data['sortby']
        sortby = ",".join(sortby)

        order = request.data['order']
        
        limit = settings.LIMIT(request.data)
        filterquery = ""
        for key,value in filterby.items():
            if type(value)==list:
                value = "', '".join(value)
                filterquery = filterquery + f" and master_position.{key} in ('{value}')"
            else:
                filterquery = filterquery + f" and master_position.{key}='{value}'"
        #print("filterquery", filterquery)
        sql = f"select master_position.* from {model_name} WHERE 1=1 {filterquery} {search} order by master_position.id desc {limit}"
        
        #for search SELECT * FROM `deliverynote_deliverynote` WHERE SalesPersonCode=18 and (U_MR_NO like "%971%" or CardCode like "%11%");
        
        # sql = f"select master_position.*, master_yearly_baseline.emp_ref_no from {model_name} left join master_yearly_baseline on master_yearly_baseline.position_id=master_position.position_id WHERE 1=1 {filterquery} order by master_position.id desc {limit}"
        #sql = select master_position.*, master_yearly_baseline.emp_ref_no from master_position left join master_yearly_baseline on  master_position.position_id = master_yearly_baseline.position_id WHERE 1=1  and master_position.country='japan' order by master_position.baseline_total desc limit 10 OFFSET 20
        #sql = select master_position.*, ifnull(master_yearly_baseline.emp_ref_no, "NA") as emp_ref_no from master_position left join master_yearly_baseline on  master_position.position_id = master_yearly_baseline.position_id WHERE 1=1  and master_position.country='India' order by master_position.baseline_total desc
        
        #print(sql)
        mycursor.execute(sql)
        objs = mycursor.fetchall()
        counts = mycursor.rowcount
        

        
        sql = f"select count(id) as count from {model_name} WHERE 1=1 {filterquery}"
        #print(sql)
        mycursor.execute(sql)
        count = mycursor.fetchone()
        #print(f'count: {count}')
        
        mycursor.close()
        
        #objs = model.objects.filter(**filterby).values(*fields).order_by("-id")
        return Response({"message":"successful","status":200, "data":objs, "counts":count})
    except Exception as e:
        return Response({"message":str(e),"status":"201","data":[]})

@api_view(['POST'])
def get_list_pos_bkp(request):
    try:		
        
        mydb = settings.CONNECT()
        mycursor = mydb.cursor(dictionary=True, buffered=True)
        #print(mycursor)

        model_name = request.data['model']
        filterby = request.data['filterby']
        fields = request.data['fields']
        if len(fields) > 0:
            fields = ",".join(fields)
        else:
            fields ="*"
        sortby = request.data['sortby']
        sortby = ",".join(sortby)

        order = request.data['order']
        
        limit = settings.LIMIT(request.data)
        filterquery = ""
        for key,value in filterby.items():
            if type(value)==list:
                value = "', '".join(value)
                filterquery = filterquery + f" and master_position.{key} in ('{value}')"
            else:
                filterquery = filterquery + f" and master_position.{key}='{value}'"
        #print("filterquery", filterquery)
        sql = f"select master_position.* from {model_name} WHERE 1=1 {filterquery} order by master_position.id desc {limit}"
        
        # sql = f"select master_position.*, master_yearly_baseline.emp_ref_no from {model_name} left join master_yearly_baseline on master_yearly_baseline.position_id=master_position.position_id WHERE 1=1 {filterquery} order by master_position.id desc {limit}"
        #sql = select master_position.*, master_yearly_baseline.emp_ref_no from master_position left join master_yearly_baseline on  master_position.position_id = master_yearly_baseline.position_id WHERE 1=1  and master_position.country='japan' order by master_position.baseline_total desc limit 10 OFFSET 20
        #sql = select master_position.*, ifnull(master_yearly_baseline.emp_ref_no, "NA") as emp_ref_no from master_position left join master_yearly_baseline on  master_position.position_id = master_yearly_baseline.position_id WHERE 1=1  and master_position.country='India' order by master_position.baseline_total desc
        
        #print(sql)
        mycursor.execute(sql)
        objs = mycursor.fetchall()
        counts = mycursor.rowcount
        

        
        sql = f"select count(id) as count from {model_name} WHERE 1=1 {filterquery}"
        #print(sql)
        mycursor.execute(sql)
        count = mycursor.fetchone()
        #print(f'count: {count}')
        
        mycursor.close()
        
        #objs = model.objects.filter(**filterby).values(*fields).order_by("-id")
        return Response({"message":"successful","status":200, "data":objs, "counts":count})
    except Exception as e:
        return Response({"message":str(e),"status":"201","data":[]})
        
#{"model":"master_position", "position_id":"ZAM_115", "child":["master_position_other_pay", "master_position_workflow"]}
@api_view(['POST'])
def get_one(request):
    try:		
        
        mydb = settings.CONNECT()
        mycursor = mydb.cursor(dictionary=True, buffered=True)
        #print(mycursor)

        model_name = request.data['model']
        child_name = request.data['child']
        position_id = request.data['position_id']
        
        sql = f"select * from {model_name} where position_id='{position_id}'"
        # sql = f"select master_position.*, master_yearly_baseline.emp_ref_no from {model_name} left join master_yearly_baseline on master_yearly_baseline.position_id=master_position.position_id WHERE master_position.position_id='{position_id}'"

        #print(sql)
        mycursor.execute(sql)
        objs = mycursor.fetchall()
        for child in child_name:
            #print(child)
            sql = f"select * from {child} where position_id='{position_id}'"
            #print(sql)
            mycursor.execute(sql)
            child_obj = mycursor.fetchall()
            #print("ch", child_obj)
            objs[0][child] = child_obj
        
        mycursor.close()
        
        return Response({"message":"successful","status":200, "data":objs})
    except Exception as e:
        return Response({"message":str(e),"status":"201","data":[]})
        
        
def get_list_pos_filter(request):
    try:		
        
        mydb = settings.CONNECT()
        mycursor = mydb.cursor(dictionary=True, buffered=True)
        #print(mycursor)

        model_name = request.data['model']
        filterby = request.data['filterby']
        if "search" in request.data:
            search = request.data['search']
            
            if len(search) >= 3:
                search = f" and (position_id like '%{search}%' or position_name like '%{search}%' or emp_ref_no like '%{search}%')"
            else:
                search =""
        else:
            search =""
        
        fields = request.data['fields']
        if len(fields) > 0:
            fields = ",".join(fields)
        else:
            fields ="*"
        sortby = request.data['sortby']
        sortby = ",".join(sortby)

        order = request.data['order']
        
        #limit = settings.LIMIT(request.data)
        filterquery = ""
        for key,value in filterby.items():
            if type(value)==list:
                value = "', '".join(value)
                filterquery = filterquery + f" and master_position.{key} in ('{value}')"
            else:
                filterquery = filterquery + f" and master_position.{key}='{value}'"
        #print("filterquery", filterquery)
        #sql = f"select master_position.* from {model_name} WHERE 1=1 {filterquery} {search} order by master_position.id desc {limit}"
        sql = f"select master_position.* from {model_name} WHERE 1=1 {filterquery} {search} order by master_position.id desc"
        
        #print(sql)
        mycursor.execute(sql)
        objs = mycursor.fetchall()
        
        mycursor.close()
        return 200, objs
    except Exception as e:
        return 201, str(e)

# Export Position Report
@api_view(['POST'])
def position_report(request):
    try:
        
        res = get_list_pos_filter(request)
        if res[0] == 200:
            
            path = "PositionReport.xlsx"
            wb_obj = openpyxl.load_workbook(path)
            fontStyle = Font(name = 'Calibri', size = "10")

            Objs = res[1]
            rowNo = 2
            for obj in Objs:   
                #print(obj)    
                """
                if OrderRequestDocumentLines.objects.filter(OrderID=obj.id).exists():
                    item_name = OrderRequestDocumentLines.objects.filter(OrderID=obj.id).values_list("ItemDescription", flat=True)
                    system = OrderRequestDocumentLines.objects.filter(OrderID=obj.id).values_list("System", flat=True)
                    system = ','.join(system)
                    item_name = ','.join(item_name)
                    #print(system)
                    
                else:
                    item_name=''
                    system = "" 
                """
                wb_obj['Sheet1'].cell(row = rowNo, column = 1, value = rowNo-1).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 2, value = obj['position_id']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 3, value = obj['emp_ref_no']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 4, value = obj['position_name']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 5, value = obj['region']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 6, value = obj['country']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 7, value = obj['baseline_total']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 8, value = obj['baseline_year']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 9, value = obj['created_date']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 10, value = obj['requester_email']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 11, value = obj['status']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 12, value = obj['regional_hr']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 13, value = obj['is_vacant']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 14, value = obj['is_active']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 15, value = obj['is_delta']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 16, value = obj['pos_grade']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 17, value = obj['pos_function']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 18, value = obj['pos_bu']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 19, value = obj['pos_fte']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 20, value = obj['pos_type']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 21, value = obj['pos_manager_emp_id']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 22, value = obj['pos_manager']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 23, value = obj['pos_manager_job_title']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 24, value = obj['pos_add_cash']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 25, value = obj['pos_bonus']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 26, value = obj['pos_add_remark']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 27, value = obj['regionalhr_comment']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 28, value = obj['approver_group_hr']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 29, value = obj['approver_group_comment']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 30, value = obj['amount_based_approval']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 31, value = obj['pos_return_invest_remark']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 32, value = obj['pos_local_curr']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 33, value = obj['pos_car_allowance']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 34, value = obj['approver_2_email']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 35, value = obj['approver_2_remark']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 36, value = obj['pos_currency']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 37, value = obj['new_annual_amount']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 38, value = obj['is_approved']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 39, value = obj['pos_car_allowance_type']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 40, value = obj['pos_end_date']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 41, value = obj['pos_revenue']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 42, value = obj['pos_ebita']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 43, value = obj['pos_ocf']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 44, value = obj['pos_netcashflow']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 45, value = obj['position_grossprofit']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 46, value = obj['position_grossmargin']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 47, value = obj['eff_from']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 48, value = obj['pos_commision_plan']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 49, value = obj['pos_cxo_tasking']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 50, value = obj['pos_gross_pay']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 51, value = obj['pos_net_pay']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 52, value = obj['pos_manager_number']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 53, value = obj['pos_price_increase']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 54, value = obj['is_revenue_generating']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 55, value = obj['end_date']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 56, value = obj['pos_contract_type']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 57, value = obj['total_cost']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 58, value = obj['payroll_burden']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 59, value = obj['plane_role_start_date']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 60, value = obj['plane_role_end_date']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 61, value = obj['plane_severance']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 62, value = obj['position_pos_bonus']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 63, value = obj['postion_baseline_total']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 64, value = obj['actual_severance']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 65, value = obj['position_pos_add_cash']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 66, value = obj['planned_increase']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 67, value = obj['actual_increase']).font = fontStyle
                wb_obj['Sheet1'].cell(row = rowNo, column = 68, value = obj['pos_annual_total_cost']).font = fontStyle


                rowNo = rowNo+1

            fileName = uuid.uuid4()
            filepath = "/static/report/PositionRequest-"+str(fileName)+".xlsx"
            wb_obj.save("bridge"+filepath)
            return Response({"message":"Successfull", "status":200, "file_path":filepath})  
        else:
            return Response({"message":"Error", "status":201, "data":[{"Error":res[1]}]})
    except Exception as e:
        #print("Exception :", str(e))        
        return Response({"message":"Error", "status":201, "data":[{"Error":str(e)}]})